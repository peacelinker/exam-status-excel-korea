"""XLSX 7개 시트 구성, 빈 셀, CSV 인코딩 테스트."""

from __future__ import annotations

import csv
from copy import deepcopy
from io import BytesIO, StringIO

import pytest
from openpyxl import load_workbook

from report_exporter import (
    AGGREGATE_HEADERS,
    PASTE_REGION_ORDER,
    SHEET_NAMES,
    create_csv_bytes,
    create_result_workbook,
)
from utils import AppError


@pytest.fixture(scope="module")
def exported_workbook(sample_result):
    return load_workbook(BytesIO(create_result_workbook(sample_result)), data_only=True)


def test_result_workbook_has_exact_sheet_order(exported_workbook):
    assert exported_workbook.sheetnames == SHEET_NAMES


def test_aggregate_column_order(exported_workbook):
    worksheet = exported_workbook["집계결과"]
    assert [worksheet.cell(1, column).value for column in range(1, 11)] == AGGREGATE_HEADERS


def test_aggregate_region_order_matches_paste_template(exported_workbook):
    worksheet = exported_workbook["집계결과"]
    assert [worksheet.cell(row, 1).value for row in range(2, 10)] == PASTE_REGION_ORDER


def test_university_row_cells_are_real_blanks(exported_workbook):
    worksheet = exported_workbook["집계결과"]
    university_row = next(row for row in range(2, worksheet.max_row + 1) if worksheet.cell(row, 1).value == "대학")
    assert all(worksheet.cell(university_row, column).value is None for column in range(2, 11))


def test_total_row_contains_verified_total(exported_workbook, sample_result):
    worksheet = exported_workbook["집계결과"]
    assert worksheet.cell(worksheet.max_row, 1).value == "전체"
    assert worksheet.cell(worksheet.max_row, 8).value == sample_result.total_counts.total_exam


def test_target_and_ratio_total_cells_are_blank(exported_workbook):
    worksheet = exported_workbook["집계결과"]
    assert worksheet.cell(worksheet.max_row, 7).value is None
    assert worksheet.cell(worksheet.max_row, 10).value is None


def test_aggregate_sheet_is_plain_paste_range_without_filter(exported_workbook):
    worksheet = exported_workbook["집계결과"]
    assert worksheet.max_row == 10
    assert worksheet.max_column == 10
    assert worksheet.freeze_panes is None
    assert worksheet.auto_filter.ref is None


def test_missing_fixed_regions_are_zero_filled(exported_workbook):
    worksheet = exported_workbook["집계결과"]
    hapjeong_row = next(row for row in range(2, 10) if worksheet.cell(row, 1).value == "합정")
    assert [worksheet.cell(hapjeong_row, column).value for column in range(2, 7)] == [0, 0, 0, 0, 0]
    assert worksheet.cell(hapjeong_row, 8).value == 0
    assert worksheet.cell(hapjeong_row, 9).value == 0


def test_template_header_and_total_fills(exported_workbook):
    worksheet = exported_workbook["집계결과"]
    assert worksheet["A1"].fill.fgColor.rgb.endswith("DDEBF7")
    assert worksheet["A10"].fill.fgColor.rgb.endswith("FFF200")


def test_total_sheet_contains_absence_breakdown(exported_workbook, sample_result):
    worksheet = exported_workbook["전체합계"]
    values = {worksheet.cell(row, 1).value: worksheet.cell(row, 2).value for row in range(3, worksheet.max_row + 1)}
    assert values["명시적 미응시 합계"] == sample_result.total_counts.explicit_absent
    assert values["시험현황 공란 합계"] == sample_result.total_counts.blank_status
    assert values["최종 미응시자 합계"] == sample_result.total_counts.absent_total


def test_status_validation_regular_total_reconciles(exported_workbook):
    worksheet = exported_workbook["상태값검산"]
    for row in range(3, worksheet.max_row + 1):
        assert worksheet.cell(row, 4).value == worksheet.cell(row, 2).value + worksheet.cell(row, 3).value


def test_header_sheet_has_expected_columns(exported_workbook):
    worksheet = exported_workbook["헤더확인"]
    assert [worksheet.cell(2, column).value for column in range(1, 7)] == [
        "시트명", "열", "기대값", "실제값", "판정", "설명"
    ]


def test_special_items_sheet_records_selected_sheet(exported_workbook):
    worksheet = exported_workbook["특이사항"]
    categories = [worksheet.cell(row, 1).value for row in range(3, worksheet.max_row + 1)]
    assert "선택한 분석 시트" in categories


def test_excluded_rows_sheet_records_university(exported_workbook):
    worksheet = exported_workbook["제외행"]
    reasons = [worksheet.cell(row, 6).value for row in range(3, worksheet.max_row + 1)]
    assert "대학 지역" in reasons


def test_analysis_info_does_not_copy_individual_names(exported_workbook):
    worksheet = exported_workbook["분석정보"]
    all_values = " ".join(str(cell.value) for row in worksheet.iter_rows() for cell in row if cell.value is not None)
    assert "익명-01" not in all_values


def test_generated_workbook_can_be_reopened(sample_result):
    data = create_result_workbook(sample_result)
    reopened = load_workbook(BytesIO(data), data_only=True)
    assert reopened["집계결과"]["A1"].value


def test_csv_has_utf8_bom(sample_result):
    assert create_csv_bytes(sample_result).startswith(b"\xef\xbb\xbf")


def test_csv_header_order(sample_result):
    text = create_csv_bytes(sample_result).decode("utf-8-sig")
    rows = list(csv.reader(StringIO(text)))
    assert rows[0] == AGGREGATE_HEADERS


def test_csv_university_numeric_fields_are_empty(sample_result):
    text = create_csv_bytes(sample_result).decode("utf-8-sig")
    rows = list(csv.DictReader(StringIO(text)))
    university = next(row for row in rows if row["지역"] == "대학")
    assert all(university[column] == "" for column in AGGREGATE_HEADERS[1:])


def test_csv_uses_fixed_region_order_and_total_label(sample_result):
    text = create_csv_bytes(sample_result).decode("utf-8-sig")
    rows = list(csv.DictReader(StringIO(text)))
    assert [row["지역"] for row in rows[:-1]] == PASTE_REGION_ORDER
    assert rows[-1]["지역"] == "전체"


def test_validation_failure_blocks_xlsx(sample_result):
    invalid = deepcopy(sample_result)
    invalid.validation_differences.append(type("Difference", (), {})())
    with pytest.raises(AppError, match="검산"):
        create_result_workbook(invalid)


def test_validation_failure_blocks_csv(sample_result):
    invalid = deepcopy(sample_result)
    invalid.validation_differences.append(type("Difference", (), {})())
    with pytest.raises(AppError, match="검산"):
        create_csv_bytes(invalid)
