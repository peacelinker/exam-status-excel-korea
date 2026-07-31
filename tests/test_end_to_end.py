"""업로드 바이트부터 결과 재열기까지의 종단 테스트."""

from __future__ import annotations

import hashlib
from io import BytesIO

from openpyxl import load_workbook

from excel_analyzer import analyze_selected_sheet
from report_exporter import create_csv_bytes, create_result_workbook
from sheet_detector import discover_candidate_sheets


def test_end_to_end_pipeline(sample_workbook_bytes):
    candidates = discover_candidate_sheets(sample_workbook_bytes)
    assert any(item.name == "0719" and item.is_analyzable for item in candidates)
    result = analyze_selected_sheet(sample_workbook_bytes, "0719", source_filename="익명샘플.xlsx")
    assert result.validation_passed
    xlsx_bytes = create_result_workbook(result)
    csv_bytes = create_csv_bytes(result)
    assert load_workbook(BytesIO(xlsx_bytes), data_only=True).sheetnames[0] == "집계결과"
    assert csv_bytes.startswith(b"\xef\xbb\xbf")


def test_original_workbook_bytes_are_unchanged(sample_workbook_bytes):
    before = hashlib.sha256(sample_workbook_bytes).hexdigest()
    result = analyze_selected_sheet(sample_workbook_bytes, "0719")
    create_result_workbook(result)
    create_csv_bytes(result)
    after = hashlib.sha256(sample_workbook_bytes).hexdigest()
    assert before == after


def test_output_values_come_from_sample_cells(sample_workbook_bytes):
    result = analyze_selected_sheet(sample_workbook_bytes, "0719")
    rows = {row["지역"]: row for row in result.aggregate_rows(include_total=False)}
    assert rows["서대문"]["전체응시"] == 6
    assert rows["마포"]["전체응시"] == 4
    assert rows["신촌"]["전체응시"] == 1


def test_output_workbook_is_distinct_from_source(sample_workbook_bytes):
    result = analyze_selected_sheet(sample_workbook_bytes, "0719")
    output = create_result_workbook(result)
    assert hashlib.sha256(output).digest() != hashlib.sha256(sample_workbook_bytes).digest()


def test_generated_output_can_be_processed_twice(sample_workbook_bytes):
    first = analyze_selected_sheet(sample_workbook_bytes, "0719")
    first_output = create_result_workbook(first)
    second_output = create_result_workbook(first)
    assert load_workbook(BytesIO(first_output))["전체합계"]["A1"].value
    assert load_workbook(BytesIO(second_output))["분석정보"]["A1"].value
