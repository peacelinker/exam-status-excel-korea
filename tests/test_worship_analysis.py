"""구역예배 A·D·H 실제 셀 집계와 재적 공란 처리 테스트."""

from __future__ import annotations

import hashlib
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from utils import AppError
from worship_analyzer import analyze_worship_sheet, discover_worship_sheets
from worship_exporter import create_worship_csv, create_worship_workbook
from worship_models import WORSHIP_REGIONS


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    guide = workbook.active
    guide.title = "안내"
    guide["A1"] = "안내"

    worksheet = workbook.create_sheet("7월4주")
    worksheet["A1"] = "지역"
    worksheet["D1"] = "이름"
    worksheet["H1"] = "참여방법"
    rows = [
        ("서대문", "익명01", "대면모임"),
        ("서대문", "익명02", "줌"),
        ("서대문", "익명03", "통화"),
        ("서대문", "익명04", " 대면모임 "),
        ("마포", "익명05", "줌"),
        ("마포", "", "대면모임"),
        ("합정", "익명06", ""),
        ("대학", "익명07", "통화"),
        ("새신", "익명08", "대면 모임"),
        ("신촌", "익명09", "통화"),
        ("홍대", "익명10", "줌"),
        ("소성", "익명11", "대면모임"),
        (None, "익명12", "줌"),
        (None, None, None),
    ]
    for row_number, (region, name, attendance) in enumerate(rows, start=2):
        worksheet.cell(row_number, 1, region)
        worksheet.cell(row_number, 4, name)
        worksheet.cell(row_number, 8, attendance)
    worksheet.row_dimensions[12].hidden = True
    worksheet.auto_filter.ref = "A1:H15"

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


@pytest.fixture()
def worship_bytes() -> bytes:
    return _workbook_bytes()


@pytest.fixture()
def worship_result(worship_bytes):
    return analyze_worship_sheet(
        worship_bytes,
        "7월4주",
        rosters={
            "서대문": 10,
            "마포": None,
            "합정": 2,
            "새신": 1,
            "신촌": 5,
            "홍대": 5,
            "소성": 5,
        },
        report_title="7/20-7/26 구역예배 성인",
        source_filename="익명구역예배.xlsx",
    )


def test_discovers_only_sheet_with_name_rows(worship_bytes):
    candidates = discover_worship_sheets(worship_bytes)
    by_name = {item.name: item for item in candidates}
    assert not by_name["안내"].is_analyzable
    assert by_name["7월4주"].is_analyzable
    assert by_name["7월4주"].recommended


def test_counts_actual_a_d_h_cells_only(worship_result):
    by_region = {item.region: item for item in worship_result.region_results}
    assert list(by_region) == WORSHIP_REGIONS
    assert by_region["서대문"].counts.validation_values() == {
        "대면": 2,
        "줌": 1,
        "전화": 1,
        "전체": 4,
    }
    assert by_region["마포"].counts.zoom == 1
    assert by_region["새신"].counts.total == 0
    assert worship_result.total_counts.validation_values() == {
        "대면": 3,
        "줌": 3,
        "전화": 2,
        "전체": 8,
    }


def test_blank_roster_keeps_counts_and_blanks_percentages(worship_result):
    by_region = {item.region: item for item in worship_result.region_results}
    mapo = by_region["마포"]
    assert mapo.counts.zoom == 1
    assert mapo.zoom_percent is None
    assert mapo.absent is None
    assert worship_result.total_roster is None
    assert worship_result.total_attendance_percent is None
    assert worship_result.total_absent is None


def test_known_roster_calculates_region_percentages(worship_result):
    seodaemun = worship_result.region_results[0]
    assert seodaemun.face_percent == pytest.approx(0.2)
    assert seodaemun.zoom_percent == pytest.approx(0.1)
    assert seodaemun.phone_percent == pytest.approx(0.1)
    assert seodaemun.attendance_percent == pytest.approx(0.4)
    assert seodaemun.absent == 6


def test_unexpected_blank_and_unsupported_rows_are_audited(worship_result):
    reasons = {item.reason for item in worship_result.excluded_rows}
    assert "D열 이름 공란" in reasons
    assert "H열 참여방식 공란" in reasons
    assert "분석 대상 외 지역" in reasons
    assert "예상하지 못한 H열 값" in reasons
    assert "A열 지역 공란" in reasons
    assert worship_result.hidden_row_count == 1
    assert worship_result.filter_applied
    assert worship_result.validation_passed


def test_invalid_roster_is_rejected(worship_bytes):
    with pytest.raises(AppError, match="0 이상의 정수"):
        analyze_worship_sheet(worship_bytes, "7월4주", rosters={"서대문": -1})


def test_original_bytes_are_unchanged(worship_bytes):
    before = hashlib.sha256(worship_bytes).hexdigest()
    result = analyze_worship_sheet(worship_bytes, "7월4주")
    create_worship_workbook(result)
    create_worship_csv(result)
    assert hashlib.sha256(worship_bytes).hexdigest() == before


def test_output_workbook_matches_target_layout(worship_result):
    output = create_worship_workbook(worship_result)
    workbook = load_workbook(BytesIO(output), data_only=True)
    worksheet = workbook["구역예배결과"]
    assert worksheet["A1"].value == "7/20-7/26 구역예배 성인"
    assert "A1:J1" in {str(item) for item in worksheet.merged_cells.ranges}
    assert [worksheet.cell(2, column).value for column in range(1, 11)] == [
        "지역", "대면", "퍼센트", "줌", "퍼센트", "전화", "퍼센트", "전체", "미참여", "출결 재적대비 %"
    ]
    assert worksheet["A3"].value == "서대문"
    assert worksheet["B3"].value == 2
    assert worksheet["C3"].value == pytest.approx(0.2)
    assert worksheet["C3"].number_format == "0.0%"
    assert worksheet["A4"].value == "마포"
    assert worksheet["D4"].value == 1
    assert worksheet["E4"].value is None
    assert worksheet["I4"].value is None
    assert worksheet["J4"].value is None
    assert worksheet["A10"].value == "전체"
    assert worksheet["H10"].value == 8
    assert worksheet["J10"].value is None
    assert worksheet["A10"].fill.fgColor.rgb.endswith("FFF200")


def test_all_rosters_enable_total_percentages(worship_bytes):
    result = analyze_worship_sheet(
        worship_bytes,
        "7월4주",
        rosters={region: 10 for region in WORSHIP_REGIONS},
    )
    assert result.total_roster == 70
    assert result.total_absent == 62
    assert result.total_attendance_percent == pytest.approx(8 / 70)


def test_csv_is_bom_encoded_and_preserves_blank_percentages(worship_result):
    output = create_worship_csv(worship_result)
    assert output.startswith(b"\xef\xbb\xbf")
    text = output.decode("utf-8-sig")
    mapo = next(line for line in text.splitlines() if line.startswith("마포,"))
    assert mapo == "마포,0,,1,,0,,1,,"

