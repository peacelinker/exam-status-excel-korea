"""개인정보가 없는 메모리 XLSX 테스트 fixture."""

from __future__ import annotations

import sys
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from excel_analyzer import analyze_selected_sheet  # noqa: E402


def workbook_bytes(workbook: Workbook) -> bytes:
    """워크북을 메모리 XLSX 바이트로 직렬화한다."""

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def add_exam_sheet(
    workbook: Workbook,
    name: str,
    rows: list[tuple[object, object, object]],
    *,
    headers: tuple[object, object, object] = ("지역", "이름", "시험현황"),
):
    """A·D·I 열에 익명 시험 자료가 있는 시트를 추가한다."""

    worksheet = workbook.create_sheet(name)
    worksheet.cell(1, 1, headers[0])
    worksheet.cell(1, 4, headers[1])
    worksheet.cell(1, 9, headers[2])
    for row_number, (region, person, status) in enumerate(rows, start=2):
        worksheet.cell(row_number, 1, region)
        worksheet.cell(row_number, 4, person)
        worksheet.cell(row_number, 9, status)
    return worksheet


@pytest.fixture()
def workbook_factory():
    """테스트마다 원하는 시트 구조를 만드는 팩터리를 제공한다."""

    def factory(specs: list[tuple[str, list[tuple[object, object, object]], tuple | None]]) -> bytes:
        workbook = Workbook()
        workbook.remove(workbook.active)
        for name, rows, headers in specs:
            add_exam_sheet(
                workbook,
                name,
                rows,
                headers=headers or ("지역", "이름", "시험현황"),
            )
        return workbook_bytes(workbook)

    return factory


@pytest.fixture(scope="session")
def sample_workbook_bytes() -> bytes:
    """모든 주요 경계 사례를 포함한 익명 샘플 XLSX를 생성한다."""

    workbook = Workbook()
    workbook.active.title = "안내"
    workbook.active["A1"] = "테스트 안내"
    add_exam_sheet(workbook, "지역전체", [("서대문", "익명-A", "정규응시")])
    add_exam_sheet(workbook, "직전시험", [("서대문", "익명-B", "미응시")])
    add_exam_sheet(workbook, "지역전체 사본", [("서대문", "익명-C", "정규응시")])

    rows = [
        ("서대문", "익명-01", "정규응시"),
        ("서대문", "익명-02", "정규응시(타지파)"),
        ("서대문", "익명-03", "대면응시"),
        ("서대문", "익명-04", "일대일응시"),
        ("서대문", "익명-05", "서면응시"),
        ("서대문", "익명-06", "비공식응시"),
        ("서대문", "익명-07", "미응시"),
        ("서대문", "익명-08", None),
        ("마포", "익명-09", " 정규응시 "),
        ("마포", "익명-10", "정규 응시"),
        ("마포", "", "정규응시"),
        (None, "익명-11", "정규응시"),
        ("대학", "익명-12", "정규응시"),
        ("마포", "익명-13", "대면응시"),
        ("마포", "익명-14", "   "),
        ("마포", "익명-15", "정규응시(타지파)"),
        ("마포", 1234, "서면응시"),
        (None, None, None),
        ("신촌", "익명-16", "정규응시"),
        ("신촌", "익명-17", "미응시"),
        ("신촌", "익명-18", '=IF(1=1,"정규응시","")'),
        ("신촌", '="익명-19"', "대면응시"),
    ]
    target = add_exam_sheet(workbook, "0719", rows)
    target.row_dimensions[20].hidden = True
    target.auto_filter.ref = "A1:I23"
    add_exam_sheet(workbook, "0726", [("마포", "익명-20", "정규응시")])
    return workbook_bytes(workbook)


@pytest.fixture(scope="session")
def sample_result(sample_workbook_bytes):
    """반복 검증에 사용할 샘플 0719 분석 결과를 제공한다."""

    return analyze_selected_sheet(
        sample_workbook_bytes,
        "0719",
        source_filename="익명샘플.xlsx",
    )
