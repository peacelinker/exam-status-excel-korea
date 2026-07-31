"""시험 자료 후보 시트를 탐색하고 필수 헤더를 검증한다."""

from __future__ import annotations

import re
from datetime import date

from openpyxl.worksheet.worksheet import Worksheet

from models import HeaderCheck, SheetCandidate
from utils import open_workbook_from_bytes, safe_cell_text

HEADER_ROW = 1
FOUR_DIGIT_SHEET_PATTERN = re.compile(r"^\d{4}$")
HEADER_RULES: dict[str, tuple[str, ...]] = {
    "A": ("지역",),
    "D": ("이름", "성명", "이름구분"),
    "I": ("시험현황", "시험 현황"),
}


def _candidate_type(sheet_name: str) -> str | None:
    """시트명 규칙에 맞는 후보 유형을 반환한다."""

    if sheet_name == "지역전체":
        return "지역전체"
    if sheet_name == "직전시험":
        return "직전시험"
    if FOUR_DIGIT_SHEET_PATTERN.fullmatch(sheet_name):
        return "날짜형 숫자 시트"
    return None


def validate_candidate_headers(
    worksheet: Worksheet,
    *,
    header_row: int = HEADER_ROW,
) -> list[HeaderCheck]:
    """A·D·I 열이 읽히는지와 헤더 이름이 기대값과 일치하는지 확인한다."""

    checks: list[HeaderCheck] = []
    for column, expected_values in HEADER_RULES.items():
        column_index = ord(column) - ord("A") + 1
        if worksheet.max_column < column_index:
            checks.append(
                HeaderCheck(
                    sheet_name=worksheet.title,
                    column=column,
                    expected_values=expected_values,
                    actual_value="",
                    status="오류",
                    description="필수 열 위치가 워크시트 범위를 벗어났습니다.",
                    readable=False,
                )
            )
            continue

        actual_value = safe_cell_text(worksheet.cell(header_row, column_index).value)
        if not actual_value:
            checks.append(
                HeaderCheck(
                    sheet_name=worksheet.title,
                    column=column,
                    expected_values=expected_values,
                    actual_value="",
                    status="오류",
                    description="헤더 값을 읽을 수 없습니다.",
                    readable=False,
                )
            )
        elif actual_value in expected_values:
            checks.append(
                HeaderCheck(
                    sheet_name=worksheet.title,
                    column=column,
                    expected_values=expected_values,
                    actual_value=actual_value,
                    status="일치",
                    description="기대 헤더와 일치합니다.",
                    readable=True,
                )
            )
        else:
            checks.append(
                HeaderCheck(
                    sheet_name=worksheet.title,
                    column=column,
                    expected_values=expected_values,
                    actual_value=actual_value,
                    status="경고",
                    description="열 위치는 읽을 수 있으나 헤더 이름이 다릅니다.",
                    readable=True,
                )
            )
    return checks


def discover_candidate_sheets(
    file_bytes: bytes,
    *,
    header_row: int = HEADER_ROW,
) -> list[SheetCandidate]:
    """이름 규칙을 만족하는 모든 시트를 검사해 후보 요약을 반환한다."""

    workbook = open_workbook_from_bytes(file_bytes, data_only=True)
    candidates: list[SheetCandidate] = []
    for worksheet in workbook.worksheets:
        candidate_type = _candidate_type(worksheet.title)
        if candidate_type is None:
            continue

        header_checks = validate_candidate_headers(worksheet, header_row=header_row)
        name_rows = 0
        status_rows = 0
        blank_status_rows = 0
        for row_number in range(header_row + 1, worksheet.max_row + 1):
            name = safe_cell_text(worksheet.cell(row_number, 4).value)
            if not name:
                continue
            name_rows += 1
            status = safe_cell_text(worksheet.cell(row_number, 9).value)
            if status:
                status_rows += 1
            else:
                blank_status_rows += 1

        reasons: list[str] = []
        if not all(check.readable for check in header_checks):
            reasons.append("필수 헤더를 읽을 수 없음")
        if worksheet.max_row <= header_row:
            reasons.append("데이터 행이 없음")
        if name_rows == 0:
            reasons.append("이름이 있는 데이터 행이 없음")

        candidates.append(
            SheetCandidate(
                name=worksheet.title,
                candidate_type=candidate_type,
                max_row=worksheet.max_row,
                name_rows=name_rows,
                status_rows=status_rows,
                blank_status_rows=blank_status_rows,
                is_analyzable=not reasons,
                header_checks=header_checks,
                exclusion_reason="; ".join(reasons),
            )
        )

    recommended = recommend_latest_sheet(candidates)
    if recommended is not None:
        recommended.recommended = True
    return candidates


def _sheet_sort_key(candidate: SheetCandidate) -> tuple[int, int, str]:
    """숫자 시트를 가능한 MMDD 날짜로 해석한 추천 정렬 키를 만든다."""

    if FOUR_DIGIT_SHEET_PATTERN.fullmatch(candidate.name):
        month = int(candidate.name[:2])
        day = int(candidate.name[2:])
        try:
            parsed = date(2000, month, day)
            return (3, parsed.timetuple().tm_yday, candidate.name)
        except ValueError:
            return (2, int(candidate.name), candidate.name)
    if candidate.name == "직전시험":
        return (1, 0, candidate.name)
    return (0, 0, candidate.name)


def recommend_latest_sheet(candidates: list[SheetCandidate]) -> SheetCandidate | None:
    """분석 가능한 후보 중 최신으로 보이는 시트를 추천하되 강제 선택하지 않는다."""

    analyzable = [candidate for candidate in candidates if candidate.is_analyzable]
    if not analyzable:
        return None
    return max(analyzable, key=_sheet_sort_key)
