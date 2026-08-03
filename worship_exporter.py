"""구역예배 분석 결과를 참고 이미지와 같은 XLSX/CSV로 내보낸다."""

from __future__ import annotations

import csv
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utils import AppError
from worship_models import WorshipAnalysisResult

HEADERS = ["재적", "지역", "대면", "퍼센트", "줌", "퍼센트", "전화", "퍼센트", "전체", "미참여", "출결 재적대비 %"]
TITLE_FILL = "E5B8C3"
HEADER_FILL = "EAC5CF"
TOTAL_FILL = "FFF200"
BLACK = "000000"
WHITE = "FFFFFF"
NAVY = "16324F"
BLUE = "2563EB"
GRID = "DCE2EA"
TABLE_BORDER = Border(
    left=Side(style="thin", color=BLACK),
    right=Side(style="thin", color=BLACK),
    top=Side(style="thin", color=BLACK),
    bottom=Side(style="thin", color=BLACK),
)


def _row_values(row: dict[str, object]) -> list[object]:
    return [
        row["재적"],
        row["지역"],
        row["대면"],
        row["퍼센트"],
        row["줌"],
        row["줌 퍼센트"],
        row["전화"],
        row["전화 퍼센트"],
        row["전체"],
        row["미참여"],
        row["출결 재적대비 %"],
    ]


def _write_primary_sheet(workbook: Workbook, result: WorshipAnalysisResult) -> None:
    worksheet = workbook.active
    worksheet.title = "구역예배결과"
    worksheet.sheet_view.showGridLines = False
    worksheet.merge_cells("A1:K1")
    worksheet["A1"] = result.report_title
    worksheet["A1"].font = Font(name="맑은 고딕", size=10, bold=True, color=BLACK)
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["A1"].fill = PatternFill("solid", fgColor=TITLE_FILL)

    for column, header in enumerate(HEADERS, start=1):
        worksheet.cell(2, column, header)

    rows = result.aggregate_rows(include_total=True)
    for row_number, row in enumerate(rows, start=3):
        for column, value in enumerate(_row_values(row), start=1):
            worksheet.cell(row_number, column, value)

    for row_number in range(1, 11):
        for column in range(1, 12):
            cell = worksheet.cell(row_number, column)
            cell.font = Font(name="맑은 고딕", size=9, bold=row_number in (1, 2, 10), color=BLACK)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = TABLE_BORDER
            if row_number == 2:
                cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
            elif row_number == 10:
                cell.fill = PatternFill("solid", fgColor=TOTAL_FILL)
            if column in (4, 6, 8, 11) and row_number >= 3:
                cell.number_format = "0.0%"
            elif column != 2 and row_number >= 3:
                cell.number_format = "#,##0"

    worksheet.row_dimensions[1].height = 19
    worksheet.row_dimensions[2].height = 19
    for row_number in range(3, 11):
        worksheet.row_dimensions[row_number].height = 18
    widths = [9, 10, 8, 10, 8, 10, 8, 10, 9, 9, 15]
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width
    worksheet.freeze_panes = "A3"
    worksheet.print_area = "A1:K10"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1


def _report_title(worksheet, text: str, columns: int) -> None:
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    cell = worksheet.cell(1, 1, text)
    cell.font = Font(name="맑은 고딕", size=14, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 28
    worksheet.sheet_view.showGridLines = False


def _report_header(worksheet, headers: list[str]) -> None:
    for column, value in enumerate(headers, start=1):
        cell = worksheet.cell(2, column, value)
        cell.font = Font(name="맑은 고딕", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color=GRID))


def _autosize(worksheet, columns: int) -> None:
    for column in range(1, columns + 1):
        width = max(
            (len(str(worksheet.cell(row, column).value)) for row in range(1, worksheet.max_row + 1) if worksheet.cell(row, column).value is not None),
            default=8,
        )
        worksheet.column_dimensions[get_column_letter(column)].width = min(max(width + 3, 10), 36)


def _write_roster_sheet(workbook: Workbook, result: WorshipAnalysisResult) -> None:
    worksheet = workbook.create_sheet("재적검산")
    headers = ["재적", "지역", "대면", "줌", "전화", "전체", "미참여", "대면 %", "줌 %", "전화 %", "출결 %"]
    _report_title(worksheet, "지역별 재적 및 출결 검산", len(headers))
    _report_header(worksheet, headers)
    rows = [*result.region_results]
    for row_number, item in enumerate(rows, start=3):
        values = [
            item.roster,
            item.region,
            item.counts.face_to_face,
            item.counts.zoom,
            item.counts.phone,
            item.counts.total,
            item.absent,
            item.face_percent,
            item.zoom_percent,
            item.phone_percent,
            item.attendance_percent,
        ]
        for column, value in enumerate(values, start=1):
            worksheet.cell(row_number, column, value)
    total = result.total_row()
    total_values = [
        result.total_roster,
        "전체",
        total["대면"],
        total["줌"],
        total["전화"],
        total["전체"],
        total["미참여"],
        total["퍼센트"],
        total["줌 퍼센트"],
        total["전화 퍼센트"],
        total["출결 재적대비 %"],
    ]
    for column, value in enumerate(total_values, start=1):
        worksheet.cell(10, column, value)
    for row in range(3, 11):
        for column in range(1, 12):
            cell = worksheet.cell(row, column)
            cell.font = Font(name="맑은 고딕", size=10, bold=row == 10)
            cell.border = Border(bottom=Side(style="thin", color=GRID))
            cell.alignment = Alignment(horizontal="left" if column == 2 else "right")
            if column >= 8:
                cell.number_format = "0.0%"
        if row == 10:
            for cell in worksheet[row]:
                cell.fill = PatternFill("solid", fgColor="FFF4B8")
    worksheet.freeze_panes = "A3"
    _autosize(worksheet, len(headers))


def _write_excluded_sheet(workbook: Workbook, result: WorshipAnalysisResult) -> None:
    worksheet = workbook.create_sheet("제외행")
    headers = ["시트명", "행번호", "지역", "이름", "H열 값", "제외 사유"]
    _report_title(worksheet, "집계 제외 행", len(headers))
    _report_header(worksheet, headers)
    for row_number, item in enumerate(result.excluded_rows, start=3):
        values = [item.sheet_name, item.row_number, item.region, item.name, item.attendance, item.reason]
        for column, value in enumerate(values, start=1):
            worksheet.cell(row_number, column, value)
    worksheet.freeze_panes = "A3"
    worksheet.auto_filter.ref = f"A2:F{max(2, worksheet.max_row)}"
    _autosize(worksheet, len(headers))


def _write_info_sheet(workbook: Workbook, result: WorshipAnalysisResult) -> None:
    worksheet = workbook.create_sheet("분석정보")
    _report_title(worksheet, "구역예배 분석 실행 정보", 2)
    _report_header(worksheet, ["항목", "값"])
    info = [
        ("분석 일시", result.analyzed_at.isoformat(timespec="seconds")),
        ("원본 파일명", result.source_filename),
        ("선택한 시트명", result.selected_sheet),
        ("분석 열", "A열 지역 · D열 이름 · H열 참여방식"),
        ("재적 자동값", "A열이 대상 지역이고 D열 이름이 있는 실제 행 수"),
        ("대면 기준값", "대면모임"),
        ("줌 기준값", "줌"),
        ("전화 기준값", "통화"),
        ("데이터 범위", f"{result.data_start_row}-{result.last_data_row}행"),
        ("숨겨진 행 수", result.hidden_row_count),
        ("필터 설정 여부", "예" if result.filter_applied else "아니요"),
        ("수식 저장 결과값 없음", result.formula_without_cached_value_count),
        ("검산 결과", "일치" if result.validation_passed else "불일치"),
        ("집계 규칙 버전", result.rule_version),
    ]
    for row_number, values in enumerate(info, start=3):
        worksheet.cell(row_number, 1, values[0])
        worksheet.cell(row_number, 2, values[1])
    _autosize(worksheet, 2)


def create_worship_workbook(result: WorshipAnalysisResult) -> bytes:
    """검산에 성공한 결과를 새 XLSX 바이트로 생성한다."""

    if not result.validation_passed:
        raise AppError("구역예배 독립 검산 결과가 일치하지 않습니다.")
    try:
        workbook = Workbook()
        _write_primary_sheet(workbook, result)
        _write_roster_sheet(workbook, result)
        _write_excluded_sheet(workbook, result)
        _write_info_sheet(workbook, result)
        stream = BytesIO()
        workbook.save(stream)
        return stream.getvalue()
    except AppError:
        raise
    except Exception as exc:
        raise AppError("구역예배 결과 엑셀 생성에 실패했습니다.") from exc


def _csv_percent(value: object) -> str:
    if value is None:
        return ""
    return f"{float(value):.1%}"


def create_worship_csv(result: WorshipAnalysisResult) -> bytes:
    """이미지와 같은 열·지역 순서의 UTF-8 BOM CSV를 생성한다."""

    if not result.validation_passed:
        raise AppError("구역예배 독립 검산 결과가 일치하지 않습니다.")
    stream = StringIO(newline="")
    writer = csv.writer(stream)
    writer.writerow(HEADERS)
    for row in result.aggregate_rows(include_total=True):
        values = _row_values(row)
        for index in (3, 5, 7, 10):
            values[index] = _csv_percent(values[index])
        writer.writerow(["" if value is None else value for value in values])
    return stream.getvalue().encode("utf-8-sig")

