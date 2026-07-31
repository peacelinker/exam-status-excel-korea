"""검산을 통과한 분석 결과를 새 XLSX와 CSV 바이트로 생성한다."""

from __future__ import annotations

import csv
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models import AnalysisResult
from utils import AppError, compress_row_numbers

AGGREGATE_HEADERS = [
    "지역",
    "정규응시",
    "대면응시",
    "일대일응시",
    "서면 응시",
    "비공식 응시",
    "전체 응시목표",
    "전체응시",
    "미응시자",
    "전도 재적대비 %",
]
PASTE_REGION_ORDER = ["서대문", "마포", "합정", "새신", "신촌", "홍대", "대학", "소성"]
SHEET_NAMES = [
    "집계결과",
    "전체합계",
    "상태값검산",
    "헤더확인",
    "특이사항",
    "제외행",
    "분석정보",
]

NAVY = "16324F"
BLUE = "2563EB"
PALE_BLUE = "EAF2FF"
PALE_GRAY = "F5F7FA"
GRID = "DCE2EA"
WHITE = "FFFFFF"
SUCCESS = "DDF5E7"
WARNING = "FFF1D6"
TOTAL = "FFF4B8"
THIN_BORDER = Border(bottom=Side(style="thin", color=GRID))
PASTE_HEADER = "DDEBF7"
PASTE_TOTAL = "FFF200"
PASTE_GRID = "000000"
PASTE_BORDER = Border(
    left=Side(style="thin", color=PASTE_GRID),
    right=Side(style="thin", color=PASTE_GRID),
    top=Side(style="thin", color=PASTE_GRID),
    bottom=Side(style="thin", color=PASTE_GRID),
)


def _sheet(workbook: Workbook, name: str):
    """기존 시트를 재사용하거나 새 시트를 만든다."""

    if name in workbook.sheetnames:
        return workbook[name]
    return workbook.create_sheet(name)


def _title(worksheet, text: str, last_column: int) -> None:
    """워크시트 상단에 일관된 보고서 제목 띠를 적용한다."""

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    cell = worksheet.cell(1, 1, text)
    cell.font = Font(name="맑은 고딕", size=16, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(vertical="center", horizontal="left")
    worksheet.row_dimensions[1].height = 30
    worksheet.sheet_view.showGridLines = False


def _header(worksheet, row: int, headers: list[str]) -> None:
    """표 헤더에 배경색, 중앙 정렬, 하단 구분선을 적용한다."""

    for column, value in enumerate(headers, start=1):
        cell = worksheet.cell(row, column, value)
        cell.font = Font(name="맑은 고딕", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    worksheet.row_dimensions[row].height = 28


def _format_body(worksheet, start_row: int, end_row: int, columns: int) -> None:
    """본문의 글꼴, 정렬, 교대 배경, 숫자 형식을 정리한다."""

    for row in range(start_row, end_row + 1):
        fill = PatternFill("solid", fgColor=PALE_GRAY if row % 2 == 0 else WHITE)
        for column in range(1, columns + 1):
            cell = worksheet.cell(row, column)
            cell.font = Font(name="맑은 고딕", size=10, color="273142")
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="right" if isinstance(cell.value, (int, float)) else "left",
                vertical="center",
                wrap_text=column == 1,
            )
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"


def _autosize(worksheet, *, max_width: int = 48) -> None:
    """내용 길이에 맞춰 열 너비를 조정하되 과도한 폭은 제한한다."""

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        length = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None),
            default=8,
        )
        worksheet.column_dimensions[column_letter].width = min(max(length + 3, 10), max_width)


def _paste_ready_rows(result: AnalysisResult) -> list[dict[str, object]]:
    """두 번째 참고 이미지의 고정 지역·열 순서에 맞는 붙여넣기용 행을 만든다."""

    by_region = {item.region: item for item in result.region_results}
    unsupported = [region for region in by_region if region not in PASTE_REGION_ORDER]
    if unsupported:
        raise AppError(
            "붙여넣기 양식에 없는 지역이 있습니다: " + ", ".join(unsupported)
        )

    rows: list[dict[str, object]] = []
    for region in PASTE_REGION_ORDER:
        region_result = by_region.get(region)
        if region == "대학" or (region_result and region_result.is_manual_region):
            values = [region, None, None, None, None, None, None, None, None, None]
        else:
            counts = region_result.counts if region_result else None
            values = [
                region,
                counts.regular_total if counts else 0,
                counts.face_to_face if counts else 0,
                counts.one_to_one if counts else 0,
                counts.written if counts else 0,
                counts.informal if counts else 0,
                None,
                counts.total_exam if counts else 0,
                counts.absent_total if counts else 0,
                None,
            ]
        rows.append(dict(zip(AGGREGATE_HEADERS, values)))
    return rows


def _paste_ready_total_row(result: AnalysisResult) -> dict[str, object]:
    """붙여넣기 양식의 마지막 전체 행을 만든다."""

    counts = result.total_counts
    values = [
        "전체",
        counts.regular_total,
        counts.face_to_face,
        counts.one_to_one,
        counts.written,
        counts.informal,
        None,
        counts.total_exam,
        counts.absent_total,
        None,
    ]
    return dict(zip(AGGREGATE_HEADERS, values))


def write_aggregate_sheet(workbook: Workbook, result: AnalysisResult) -> None:
    """참고 양식과 같은 A1:J10 붙여넣기용 결과 표를 작성한다."""

    worksheet = _sheet(workbook, "집계결과")
    worksheet.sheet_view.showGridLines = False

    for column, value in enumerate(AGGREGATE_HEADERS, start=1):
        worksheet.cell(1, column, value)

    rows = _paste_ready_rows(result)
    rows.append(_paste_ready_total_row(result))
    for row_number, row in enumerate(rows, start=2):
        for column, header in enumerate(AGGREGATE_HEADERS, start=1):
            worksheet.cell(row_number, column, row[header])

    for row in range(1, 11):
        for column in range(1, 11):
            cell = worksheet.cell(row, column)
            cell.font = Font(name="맑은 고딕", size=9, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = PASTE_BORDER
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

    for cell in worksheet[1]:
        cell.fill = PatternFill("solid", fgColor=PASTE_HEADER)
    for cell in worksheet[10]:
        cell.fill = PatternFill("solid", fgColor=PASTE_TOTAL)

    worksheet["A8"].comment = Comment(
        "대학은 수기 입력용 공란 행이며 자동 집계와 전체 합계에서 제외됩니다.",
        "자동 집계기",
    )
    worksheet.row_dimensions[1].height = 19
    for row in range(2, 11):
        worksheet.row_dimensions[row].height = 18
    widths = [11, 11, 11, 12, 11, 12, 14, 11, 11, 15]
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width
    worksheet.print_area = "A1:J10"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1


def write_total_sheet(workbook: Workbook, result: AnalysisResult) -> None:
    """전체 합계와 제외·경고 건수를 세로형 요약 표로 작성한다."""

    worksheet = _sheet(workbook, "전체합계")
    _title(worksheet, "전체 집계 요약", 2)
    _header(worksheet, 2, ["항목", "값"])
    metrics = [
        ("정규응시 합계", result.total_counts.regular_total),
        ("대면응시 합계", result.total_counts.face_to_face),
        ("일대일응시 합계", result.total_counts.one_to_one),
        ("서면응시 합계", result.total_counts.written),
        ("비공식응시 합계", result.total_counts.informal),
        ("전체응시 합계", result.total_counts.total_exam),
        ("명시적 미응시 합계", result.total_counts.explicit_absent),
        ("시험현황 공란 합계", result.total_counts.blank_status),
        ("최종 미응시자 합계", result.total_counts.absent_total),
        ("예상하지 못한 상태값 행 수", result.total_counts.unexpected),
        (
            "대학 지역 제외 행 수",
            sum(row.reason == "대학 지역" for row in result.excluded_rows),
        ),
        (
            "지역 공란 제외 행 수",
            sum(row.reason == "지역 공란" for row in result.excluded_rows),
        ),
    ]
    for row, values in enumerate(metrics, start=3):
        worksheet.cell(row, 1, values[0])
        worksheet.cell(row, 2, values[1])
    _format_body(worksheet, 3, 2 + len(metrics), 2)
    worksheet.freeze_panes = "A3"
    worksheet.auto_filter.ref = f"A2:B{2 + len(metrics)}"
    _autosize(worksheet)


def write_status_validation_sheet(workbook: Workbook, result: AnalysisResult) -> None:
    """각 지역에서 발견된 원시 상태값과 파생 합계를 작성한다."""

    worksheet = _sheet(workbook, "상태값검산")
    headers = [
        "지역",
        "정규응시",
        "정규응시(타지파)",
        "정규응시 합계",
        "대면응시",
        "일대일응시",
        "서면응시",
        "비공식응시",
        "명시적 미응시",
        "시험현황 공란",
        "예상하지 못한 값",
        "전체 데이터 행",
    ]
    _title(worksheet, "상태값별 검산", len(headers))
    _header(worksheet, 2, headers)
    row_number = 3
    for region_result in result.region_results:
        if region_result.is_manual_region:
            continue
        counts = region_result.counts
        values = [
            region_result.region,
            counts.regular_standard,
            counts.regular_other_tribe,
            counts.regular_total,
            counts.face_to_face,
            counts.one_to_one,
            counts.written,
            counts.informal,
            counts.explicit_absent,
            counts.blank_status,
            counts.unexpected,
            counts.total_exam
            + counts.explicit_absent
            + counts.blank_status
            + counts.unexpected,
        ]
        for column, value in enumerate(values, start=1):
            worksheet.cell(row_number, column, value)
        row_number += 1
    if row_number > 3:
        _format_body(worksheet, 3, row_number - 1, len(headers))
    worksheet.freeze_panes = "A3"
    worksheet.auto_filter.ref = f"A2:L{max(2, row_number - 1)}"
    _autosize(worksheet, max_width=24)


def write_header_sheet(workbook: Workbook, result: AnalysisResult) -> None:
    """헤더 기대값과 실제값, 판정, 설명을 작성한다."""

    worksheet = _sheet(workbook, "헤더확인")
    headers = ["시트명", "열", "기대값", "실제값", "판정", "설명"]
    _title(worksheet, "필수 헤더 확인 결과", len(headers))
    _header(worksheet, 2, headers)
    for row, check in enumerate(result.header_checks, start=3):
        values = [
            check.sheet_name,
            check.column,
            check.expected_label,
            check.actual_value,
            check.status,
            check.description,
        ]
        for column, value in enumerate(values, start=1):
            worksheet.cell(row, column, value)
    _format_body(worksheet, 3, 2 + len(result.header_checks), len(headers))
    for row, check in enumerate(result.header_checks, start=3):
        worksheet.cell(row, 5).fill = PatternFill(
            "solid", fgColor=SUCCESS if check.status == "일치" else WARNING
        )
    worksheet.freeze_panes = "A3"
    worksheet.auto_filter.ref = f"A2:F{2 + len(result.header_checks)}"
    _autosize(worksheet)


def write_special_items_sheet(workbook: Workbook, result: AnalysisResult) -> None:
    """공란, 예상 밖 상태, 구조 경고 등 특이사항을 작성한다."""

    worksheet = _sheet(workbook, "특이사항")
    headers = ["구분", "값", "건수", "행번호 범위", "설명"]
    _title(worksheet, "특이사항 및 분석 경고", len(headers))
    _header(worksheet, 2, headers)
    for row, item in enumerate(result.special_items, start=3):
        values = [item.category, item.value, item.count, item.row_range, item.description]
        for column, value in enumerate(values, start=1):
            worksheet.cell(row, column, value)
    end_row = 2 + len(result.special_items)
    if result.special_items:
        _format_body(worksheet, 3, end_row, len(headers))
    worksheet.freeze_panes = "A3"
    worksheet.auto_filter.ref = f"A2:E{max(2, end_row)}"
    _autosize(worksheet)


def write_excluded_rows_sheet(workbook: Workbook, result: AnalysisResult) -> None:
    """집계에서 제외된 행의 위치와 사유를 감사 표로 작성한다."""

    worksheet = _sheet(workbook, "제외행")
    headers = ["시트명", "행번호", "지역", "이름", "시험현황", "제외 사유"]
    _title(worksheet, "집계 제외 행", len(headers))
    _header(worksheet, 2, headers)
    for row, item in enumerate(result.excluded_rows, start=3):
        values = [
            item.sheet_name,
            item.row_number,
            item.region,
            item.name,
            item.status,
            item.reason,
        ]
        for column, value in enumerate(values, start=1):
            worksheet.cell(row, column, value)
    end_row = 2 + len(result.excluded_rows)
    if result.excluded_rows:
        _format_body(worksheet, 3, end_row, len(headers))
    worksheet.freeze_panes = "A3"
    worksheet.auto_filter.ref = f"A2:F{max(2, end_row)}"
    _autosize(worksheet)


def write_analysis_info_sheet(workbook: Workbook, result: AnalysisResult) -> None:
    """개인 이름 없이 파일·시트·분석 범위·검산 정보를 작성한다."""

    worksheet = _sheet(workbook, "분석정보")
    _title(worksheet, "분석 실행 정보", 2)
    _header(worksheet, 2, ["항목", "값"])
    info = [
        ("분석 일시", result.analyzed_at.isoformat(timespec="seconds")),
        ("원본 파일명", result.source_filename),
        ("선택한 시트명", result.selected_sheet),
        ("후보 시트 목록", ", ".join(result.candidate_names)),
        ("사용한 헤더 행", result.header_row),
        ("데이터 시작 행", result.data_start_row),
        ("마지막 데이터 행", result.last_data_row),
        ("숨겨진 행 수", result.hidden_row_count),
        ("필터 설정 여부", "예" if result.filter_applied else "아니요"),
        (
            "검산 결과",
            "일치" if result.validation_passed else "불일치",
        ),
        ("집계 규칙 버전", result.rule_version),
        (
            "수식 저장 결과값 없음",
            result.formula_without_cached_value_count,
        ),
    ]
    for row, values in enumerate(info, start=3):
        worksheet.cell(row, 1, values[0])
        worksheet.cell(row, 2, values[1])
    _format_body(worksheet, 3, 2 + len(info), 2)
    _autosize(worksheet)


def create_result_workbook(result: AnalysisResult) -> bytes:
    """검산 성공 결과를 7개 시트의 새 XLSX 바이트로 생성한다."""

    if not result.validation_passed:
        raise AppError("두 가지 검산 결과가 일치하지 않습니다.")
    try:
        workbook = Workbook()
        workbook.active.title = "집계결과"
        write_aggregate_sheet(workbook, result)
        write_total_sheet(workbook, result)
        write_status_validation_sheet(workbook, result)
        write_header_sheet(workbook, result)
        write_special_items_sheet(workbook, result)
        write_excluded_rows_sheet(workbook, result)
        write_analysis_info_sheet(workbook, result)
        stream = BytesIO()
        workbook.save(stream)
        return stream.getvalue()
    except AppError:
        raise
    except Exception as exc:
        raise AppError("결과 엑셀 생성에 실패했습니다.") from exc


def create_csv_bytes(result: AnalysisResult) -> bytes:
    """붙여넣기용 집계결과와 같은 순서의 UTF-8 BOM CSV 바이트를 생성한다."""

    if not result.validation_passed:
        raise AppError("두 가지 검산 결과가 일치하지 않습니다.")
    stream = StringIO(newline="")
    writer = csv.DictWriter(stream, fieldnames=AGGREGATE_HEADERS, extrasaction="ignore")
    writer.writeheader()
    for row in [*_paste_ready_rows(result), _paste_ready_total_row(result)]:
        writer.writerow({key: "" if value is None else value for key, value in row.items()})
    return stream.getvalue().encode("utf-8-sig")
